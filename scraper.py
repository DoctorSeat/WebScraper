import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from lxml import html
import time
from openpyxl import load_workbook
import re

PATH = "C:\\Program Files (x86)\\chromedriver.exe"
excelPath = 'C:\\Users\\Awar Satar\\Desktop\\Coding\\WebScraper\\renamethislater.xlsx'
paths = [
    '//*[@id="dkv_portrait"]/div/div[1]/section/h1/text()',  # name
    '//*[@id="dkv_portrait"]/div/div[2]/section[1]/div[1]/div[1]/p[1]/text()[2]',  # postleitzahl
    '//*[@id="collapseBasicInfos"]/div/ul[1]/li[1]/text()',  # anz. betten
    '//*[@id="collapseBasicInfos"]/div/ul[1]/li[2]/text()',  # anz. fachabteilungen
    '//*[@id="collapseGeneral"]/div/ul[3]/li/ul/li[1]/a/text()',
    # anz. ärzte //*[@id="collapseGeneral"]/div/ul[2]/li/ul/li[1]/a
    '//*[@id="collapseGeneral"]/div/ul[3]/li/ul/li[2]/a/text()',  # anz. pflege
    '//*[@id="collapseGeneral"]/div/ul[3]/li/ul/li[3]/a/text()',  # anz. therapeutisches
    '//*[@id="collapseBasicInfos"]/div/ul[3]/li[2]/text()',  # trägerart
]
paths2 = [
    '//*[@id="collapseGeneral"]/div/ul[2]/li/ul/li[1]/a/text()',  # anz. ärzte
    '//*[@id="collapseGeneral"]/div/ul[2]/li/ul/li[2]/a/text()',  # anz. pflege
    '//*[@id="collapseGeneral"]/div/ul[2]/li/ul/li[3]/a/text()',  # anz. therapeutisches
]
headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/50.0.2661.102 Safari/537.36'
}


def scrape(links):
    global string
    workbook = load_workbook(excelPath)
    sheet = workbook.active
    next_row = sheet.max_row + 1
    print("Scraping links...")
    for k in range(len(links)):
        print(links[k] + f" scraped; link no.: {k + 1}")
        response = requests.get(links[k], headers=headers)
        source_code = html.fromstring(response.content)
        for m in range(len(paths)):
            try:
                tree = source_code.xpath(paths[m])
                string = tree[0]
            except IndexError:
                try:
                    print("IndexError ... Attempting path change...")
                    tree = source_code.xpath(paths2[m - 4])
                    string = tree[0]
                except IndexError:
                    string = "XPATH ERROR"
                    print(string)
            finally:
                if ':' in string:
                    pattern = re.compile('(?<=:).*')
                    string = pattern.search(string).group()
                cell = sheet.cell(row=next_row + k, column=m + 1)
                cell.value = string

    workbook.save(excelPath)


if __name__ == "__main__":
    driver = webdriver.Chrome(PATH)
    driver.get("https://www.deutsches-krankenhaus-verzeichnis.de/app/suche/erweitert")
    select = Select(driver.find_element(By.ID, "search_whereRange"))
    select.select_by_value('1000')
    button = driver.find_element(By.ID, 'search_send')
    button.click()
    sortButton = driver.find_element(By.XPATH, '//*[@id="js_orderBetten"]/a[2]')
    sortButton.click()
    for i in range(116):
        time.sleep(5)
        linkTable = []
        table = driver.find_element(By.XPATH, '//*[@id="dkv_result_table_row"]')
        print("Fetching links....")
        for j in range(20):
            tableEntry = driver.find_element(By.XPATH, f'//*[@id="dkv_result_table_row"]/tr[{j + 1}]/td[1]')
            link = tableEntry.find_element(By.TAG_NAME, 'a').get_attribute('href')
            linkTable.append(link)
            print(link + f" on page {i + 1} entry {j + 1}")
        scrape(linkTable)
        pagination = driver.find_elements(By.TAG_NAME, 'li')
        for element in pagination:
            if "Weiter" in element.text:
                continueButton = element.find_element(By.TAG_NAME, 'a')
                print("Visiting next page...")
                continueButton.click()
        print("Taking a deep breath..")
        time.sleep(7)
